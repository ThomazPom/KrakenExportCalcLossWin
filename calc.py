from functools import reduce

import pandas as pd, pprint as pp
from openpyxl.utils.cell import get_column_letter

trades = pd.read_csv("trades.tom.csv")

trades.reindex(index=trades.index[::-1])

balances = {

}
cessions = [

]
for tradei in trades.iterrows():
    trade = tradei[1]

    if not trade["cost"] or "EUR" not in trade["pair"]:
        continue
    balances.setdefault(trade["pair"], {"buys": []})
    balance = balances.get(trade["pair"])

    buys = balance.get("buys")

    balance_before = sum([x["meta"]["vol"] for x in buys])

    result = dict(

        date=trade["time"],
        type=trade["type"],
        pair=trade["pair"],
        buy_price=trade["price"],
        total=trade["cost"],
        diff_per_unit=0,
        loss_win=0,
        balance_before=balance_before,
        vol=trade["vol"],
        balance_after=balance_before + trade["vol"],
        included_fee=trade["fee"],
        id=trade["ordertxid"],
    )

    if trade["type"] == "buy":
        buys.append({"trade": trade, "meta": {
            "price": trade["price"],
            "vol": trade["vol"],
        }})
    elif trade["type"] == "sell":

        funds_left = trade["vol"]
        coef = vols = 0
        for buy in buys.copy():
            if funds_left > 0:
                funds_left -= buy["meta"]["vol"]
                coef += buy["meta"]["vol"] * buy["meta"]["price"]
                vols += buy["meta"]["vol"]
                buys.remove(buy)
            if funds_left < 0:
                vols -= abs(funds_left)
                coef -= abs(funds_left) * buy["meta"]["price"]
                buy["meta"]["vol"] = abs(funds_left)
                buys.insert(0, buy)
                break

        avg = coef / vols if vols else 1
        diff = trade["price"] - avg
        losswin = diff * trade["vol"]
        balance_after = sum([x["meta"]["vol"] for x in buys])
        result.update(dict(
            buy_price=avg,
            diff_per_unit=diff,
            loss_win=losswin,
            balance_after=balance_after,
        ))
    cessions.append(result)

cessions.insert(0, {key: key for key in cessions[0].keys()})
import xlsxwriter

# Create an new Excel file and add a worksheet.
workbook = xlsxwriter.Workbook('result.xlsx')
worksheet = workbook.add_worksheet()
worksheet.set_column('A:Z', 12)

rowdecal = 1
worksheet.freeze_panes(1 + rowdecal, 0)
worksheet.autofilter(0 + rowdecal, 0, len(cessions) + rowdecal, len(cessions[0].keys()))

light_salmon = workbook.add_format({'bg_color': '#ffa07a'})
light_green = workbook.add_format({'bg_color': '#90ee90'})
light_blue = workbook.add_format({'color': "#FFFFFF", 'bg_color': '#4169E1'})

worksheet.set_row(0, cell_format=light_blue)

for idx1, val1 in enumerate(cessions):
    idx1 += rowdecal
    worksheet.set_row(idx1, cell_format=light_green if val1["type"] == "buy" else light_salmon if val1[
                                                                                                      "type"] == "sell" else light_blue)
    for idx2, val2 in enumerate(val1.values()):
        if type(val2) is float:
            worksheet.write_number(idx1, idx2, round(val2, 2))
            worksheet.write_formula(0, idx2,
                                    f"=SUBTOTAL(9,{get_column_letter(idx2+1)}{rowdecal + 2}:{get_column_letter(idx2+1)}{len(cessions) + rowdecal + 2})")
        else:
            worksheet.write_string(idx1, idx2, str(val2))

workbook.close()
