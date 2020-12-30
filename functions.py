import datetime, math, copy
import json
import pathlib

import requests
import pprint, os

pp = pprint.pprint
server = open("server.conf", "r").read()
not_in_list = []
qatcache = {}

import xlsxwriter

from openpyxl.utils.cell import get_column_letter


def query_all_kraken(type, typelow, user, day_zero, sleep, extra=dict(), sort="time"):
    import python3krakenex.krakenex.api as kapi_m

    k = kapi_m.API()

    k.load_key(f'{user}.key')

    continue_crawl = True
    offset = 0
    items = []
    cahename = f"{typelow}.{user}.json"
    if os.path.exists(cahename) and pathlib.Path(
            cahename).stat().st_mtime > datetime.datetime.utcnow().timestamp() - 3600 * 24:
        items = json.load(open(f"{typelow}.{user}.json", "r"))
        continue_crawl = False
    import time

    while continue_crawl:
        pp(f"query {type}")
        crawl_part = k.query_private(type, {**extra, **{'start': day_zero.timestamp(),
                                                        "end": datetime.datetime.now().timestamp(), "ofs": offset}})
        time.sleep(sleep)

        for key, val in crawl_part.get("result").get(typelow).items():
            val[f"{typelow}_id"] = key
            items.append(val)

        continue_crawl = len(crawl_part.get("result").get(typelow).keys()) == 50
        offset = offset + 50

        print(f'{offset}/{crawl_part.get("result").get("count")}')

    items.sort(key=lambda x: x.get(sort))
    json.dump(items, open(f"{typelow}.{user}.json", "w"), indent=4)
    return items


def update_balances_values(balances, item, user, timekey="time", offset=0):
    for key, val in balances.items():
        if key is not "ZEUR":
            ticker = query_at(key, item.get(timekey) + offset, user,
                              "cryptowatch-data-" + datetime.datetime.fromtimestamp(
                                  item.get(timekey) + offset).isoformat()[0:7], register_not_in_list=False)
            if not ticker:
                ticker = query_at(key, item.get(timekey) + offset, user)  # Search further ..
            if not ticker:
                continue
            val["price_atm"] = ticker.get("price")
            # pp(ticker)
            val["value_atm"] = ticker.get("price") * val["balance"]


def workbook(things_to_export, filename="export.xlsx", stylecond=lambda x: {'bg_color': '#e8e8e8'}):
    # Create an new Excel file and add a worksheet.
    workbook = xlsxwriter.Workbook(filename)

    worksheet = workbook.add_worksheet()
    worksheet.set_column('A:Z', 12)

    rowdecal = 2
    worksheet.freeze_panes(rowdecal, 0)

    light_blue = workbook.add_format({'color': "#FFFFFF", 'bg_color': '#4169E1'})
    [worksheet.set_row(i, cell_format=light_blue) for i in range(2)]
    max_seen_value_len = {

    }
    known_columns = {}
    styles = {}
    for idx1, val1 in enumerate(things_to_export):
        idx1 += rowdecal
        linestyle = stylecond(val1)
        if not styles.get(str(linestyle)):
            w_linestyle = workbook.add_format(linestyle)
            w_linestyle.set_text_wrap()
            w_linestyle.set_align('top')
            w_linestyle.set_bottom_color("DEDEDE")
            w_linestyle.set_bottom(5)
            styles[str(linestyle)] = w_linestyle

        worksheet.set_row(idx1, cell_format=styles[str(linestyle)])
        for key, val2 in val1.items():
            known_columns[key] = 1
            idx2 = list(known_columns.keys()).index(key)
            column_letter = get_column_letter(idx2 + 1)
            is_iterable = type(val2) in [list, tuple]
            is_keyval = type(val2) in [dict]
            max_seen_value_len[column_letter] = max(
                max_seen_value_len.get(column_letter, 0),
                *[len(str(x)) for x in val2] if is_iterable
                else [len(f"{key}: {val}") for key, val in val2.items()] if is_keyval
                else [len(x) for x in str(val2).split("\n")], len(key) + 3)

            if type(val2) in [float, int]:
                worksheet.write_number(idx1, idx2, round(val2, 2))
                worksheet.write_formula(0, idx2, f"=SUBTOTAL(9,{column_letter}{rowdecal + 2}:{column_letter}"
                                        + f"{len(things_to_export) + rowdecal + 2})")

            elif is_keyval:
                worksheet.write_string(idx1, idx2, "\r\n".join([f"{key}: {val}" for key, val in val2.items()]))
            elif is_iterable:
                worksheet.write_string(idx1, idx2, "\r\n".join(val2))
            else:
                worksheet.write_string(idx1, idx2, str(val2))
            worksheet.set_column(f"{column_letter}:{column_letter}", min(max_seen_value_len.get(column_letter), 50))
    for idx, key in enumerate(known_columns.keys()):
        worksheet.write_string(1, idx, key)

    worksheet.autofilter(rowdecal - 1, 0, len(things_to_export) + rowdecal, len(known_columns.keys()) - 1)
    workbook.close()


def recompose_orders(trades):
    orders = {}
    order_template = {
        "refid": None,
        "userref": None,
        "status": "closed",
        "reason": None,
        "opentm": 0,
        "closetm": math.inf,
        "starttm": 0,
        "expiretm": 0,
        "descr": {
            "pair": "XBTEUR",
            "type": "buy",
            "ordertype": "market",
            "price": "0",
            "price2": "0",
            "leverage": "none",
            "order": "buy 0.0 XBTEUR @ market",
            "close": ""
        },
        "vol": 0,
        "vol_exec": 0,
        "cost": 0,
        "fee": 0,
        "price": 0,
        "stopprice": 0,
        "limitprice": 0,
        "misc": "",
        "oflags": "fciq",
        "closed_id": ""
    }
    for trade in trades:
        order_work = orders[trade.get("ordertxid")] = orders.get(trade.get("ordertxid"), copy.deepcopy(order_template))
        order_work["a"] = 1


def search_tradebal(index, must=[], must_not=[], filter=[], should=[], minimum_should_match=0, sort=None):
    if sort is None:
        sort = [
            {"insert_date": {"order": "desc"}}
        ]
    query = {
        "query": {
            "bool": {
                "must": must,
                "filter": filter,
                "must_not": must_not,
                "should": should,
                "minimum_should_match": minimum_should_match,
                "boost": 1.0
            }
        },
        "sort": sort
    }
    r = requests.post(f"{server}/{index}/_search", headers={
        'content-type': 'application/json',
    },
                      data=json.dumps(query))

    result = r.json().get("hits")
    if not result:
        pp(r.json())
        pp(query)
    result= result.get("hits")
    if len(result) == 0:
        print("No result")
    else:
        result = result[0].get("_source")
    return result


def query_at(name, ts, user, index="cryptowatch-data-*", register_not_in_list=True):
    global qatcache, not_in_list
    if name in not_in_list:
        return False
    cahename = f"qatcache.{user}.json"
    if os.path.exists(cahename) and len(qatcache.keys()) == 0:
        qatcache = json.load(open(f"qatcache.{user}.json", "r"))
    if qatcache.get(f"{name}:{ts}"):
        return qatcache.get(f"{name}:{ts}")
    print(f"""Query {index} for balance at {datetime.datetime.fromtimestamp(
        ts).isoformat()} and {name} (utc){datetime.datetime.utcfromtimestamp(ts).isoformat()}""")
    query = {
        "query": {
            "bool": {
                "must": [
                    {
                        "exists": {
                            "field": "high"
                        }
                    },
                    {
                        "match": {
                            "name_iso": name
                        }
                    }
                    ,
                    {
                        "range": {
                            "insert_date": {
                                "gte": datetime.datetime.utcfromtimestamp(ts).isoformat()
                            }
                        }
                    }
                ]
            }
        }
        , "sort": [
            {"insert_date": {"order": "asc"}}
        ],
        "size": 1
    }

    r = requests.post(f"{server}/{index}/_search", headers={
        'content-type': 'application/json',
    },
                      data=json.dumps(query))

    result = r.json().get("hits").get("hits")
    if len(result) == 0:
        print("No result for :" + name)
        if register_not_in_list:
            not_in_list.append(name)
        return False
    else:
        result = result[0].get("_source")
    qatcache[f"{name}:{ts}"] = result
    json.dump(qatcache, open(f"qatcache.{user}.json", "w"), indent=4)
    return result
# query_at("XXBT",datetime.datetime.now().timestamp()-10,"")
